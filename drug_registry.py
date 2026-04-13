"""
drug_registry.py — 타겟 의약품 관리

유나이티드제약 Excel 파일에서 8개 의약품을 로드하고,
JSON으로 영속화하며, 새 의약품 추가를 지원한다.

사용:
    from drug_registry import DrugRegistry
    reg = DrugRegistry()
    reg.load_from_excel("path/to/excel.xlsx")
    drugs = reg.list_drugs()
    keywords = reg.generate_search_keywords(drugs[0])
"""

from __future__ import annotations

import hashlib
import json
import logging
import os
import re
import sys
import unicodedata
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

# normalizer 재사용
sys.path.insert(0, str(Path(__file__).resolve().parent / "assets" / "snippets"))
from normalizer import normalize_dosage_form, normalize_strength

logger = logging.getLogger("drug_registry")

DEFAULT_REGISTRY_PATH = Path(__file__).resolve().parent / "assets" / "drug_registry.json"


@dataclass
class TargetDrug:
    """타겟 의약품 데이터 모델.

    5개 핵심 필드: drug_type, trade_name, ingredient, strength, dosage_form
    """
    id: str                          # slug (e.g. "rosumeg-combigel")
    drug_type: str                   # 종류: 개량신약, 일반제, 항암제
    trade_name: str                  # 품목: Rosumeg Combigel
    ingredient: str                  # 성분: Rosuvastatin 5mg or 10mg + Omega-3-Acid Ethyl Esters 90 1g
    strength: str                    # 함량: 5/1000, 10/1000
    dosage_form: str                 # 제형: Cap.
    target_countries: list[str] = field(default_factory=list)  # 국가 목록
    target_regions: list[str] = field(default_factory=list)    # 구분 목록
    added_at: str = ""               # ISO timestamp
    source: str = "excel"            # "excel" or "manual"


def _slugify(text: str) -> str:
    """텍스트를 URL-safe slug로 변환."""
    text = unicodedata.normalize("NFKD", text)
    text = text.lower().strip()
    text = re.sub(r"[^a-z0-9\s-]", "", text)
    text = re.sub(r"[\s-]+", "-", text)
    return text.strip("-") or "unknown"


class DrugRegistry:
    """타겟 의약품 레지스트리.

    Excel에서 로드하거나 JSON에서 복원하고,
    새 의약품을 추가할 수 있다.
    """

    def __init__(self, registry_path: str | Path | None = None) -> None:
        self._path = Path(registry_path) if registry_path else DEFAULT_REGISTRY_PATH
        self._drugs: list[TargetDrug] = []

    # ─── 로드/저장 ─────────────────────────────────────────

    def load_from_json(self) -> list[TargetDrug]:
        """JSON 파일에서 약품 목록 로드."""
        if not self._path.exists():
            logger.info(f"레지스트리 파일 없음: {self._path}")
            return []
        with open(self._path, "r", encoding="utf-8") as f:
            data = json.load(f)
        self._drugs = [TargetDrug(**d) for d in data]
        logger.info(f"{len(self._drugs)}개 약품 로드 완료")
        return list(self._drugs)

    def save_to_json(self) -> None:
        """약품 목록을 JSON으로 저장."""
        self._path.parent.mkdir(parents=True, exist_ok=True)
        with open(self._path, "w", encoding="utf-8") as f:
            json.dump([asdict(d) for d in self._drugs], f,
                      ensure_ascii=False, indent=2)
        logger.info(f"{len(self._drugs)}개 약품 저장 → {self._path}")

    def load_from_excel(self, path: str | Path) -> list[TargetDrug]:
        """유나이티드제약 Excel 파일에서 8개 의약품 로드.

        Excel 구조:
          A: 종류, B: 품목, C: 성분, D: 함량, E: 제형, F: 구분, G: 국가
          헤더: 4행, 데이터: 5행~
          한 제품이 여러 행에 걸쳐 국가가 나열됨 (B열이 비어있으면 위 제품의 국가 추가)
        """
        import openpyxl

        path = Path(path)
        if not path.exists():
            # Unicode NFC normalization 시도
            parent = path.parent
            target_name = unicodedata.normalize("NFC", path.name)
            for f in parent.iterdir():
                if unicodedata.normalize("NFC", f.name) == target_name:
                    path = f
                    break

        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb.active

        drugs: list[TargetDrug] = []
        current_drug: TargetDrug | None = None

        for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=False):
            cells = {c.column: c.value for c in row}

            # A=1:종류, B=2:품목, C=3:성분, D=4:함량, E=5:제형, F=6:구분, G=7:국가
            drug_type = str(cells.get(1, "") or "").strip()
            trade_name = str(cells.get(2, "") or "").strip()
            ingredient = str(cells.get(3, "") or "").strip()
            strength = str(cells.get(4, "") or "").strip()
            dosage_form = str(cells.get(5, "") or "").strip()
            region = str(cells.get(6, "") or "").strip()
            country = str(cells.get(7, "") or "").strip()

            # 새 제품 행 (B열에 품목명이 있음)
            if trade_name:
                # 이전 제품 저장
                if current_drug:
                    drugs.append(current_drug)

                # 성분 줄바꿈 → 공백으로 정리
                ingredient = ingredient.replace("\n", " ").replace("  ", " ").strip()
                strength = strength.replace("\n", " ").strip()

                current_drug = TargetDrug(
                    id=_slugify(trade_name),
                    drug_type=drug_type if drug_type else (current_drug.drug_type if current_drug else ""),
                    trade_name=trade_name,
                    ingredient=ingredient,
                    strength=strength,
                    dosage_form=dosage_form,
                    target_countries=[country] if country else [],
                    target_regions=[region] if region else [],
                    added_at=datetime.now(timezone.utc).isoformat(),
                    source="excel",
                )
            elif current_drug and (country or region):
                # 기존 제품에 국가/구분 추가
                if country and country not in current_drug.target_countries:
                    current_drug.target_countries.append(country)
                if region and region not in current_drug.target_regions:
                    current_drug.target_regions.append(region)

        # 마지막 제품 저장
        if current_drug:
            drugs.append(current_drug)

        wb.close()

        self._drugs = drugs
        self.save_to_json()
        logger.info(f"Excel에서 {len(drugs)}개 약품 로드 완료")
        return list(drugs)

    # ─── CRUD ──────────────────────────────────────────────

    def list_drugs(self) -> list[TargetDrug]:
        """등록된 약품 목록 반환."""
        if not self._drugs:
            self.load_from_json()
        return list(self._drugs)

    def get_drug(self, drug_id: str) -> TargetDrug | None:
        """ID로 약품 조회."""
        for d in self.list_drugs():
            if d.id == drug_id:
                return d
        return None

    def add_drug(
        self,
        drug_type: str,
        trade_name: str,
        ingredient: str,
        strength: str,
        dosage_form: str,
        target_countries: list[str] | None = None,
    ) -> TargetDrug:
        """새 의약품 추가."""
        if not self._drugs:
            self.load_from_json()

        drug = TargetDrug(
            id=_slugify(trade_name),
            drug_type=drug_type,
            trade_name=trade_name,
            ingredient=ingredient,
            strength=strength,
            dosage_form=dosage_form,
            target_countries=target_countries or [],
            added_at=datetime.now(timezone.utc).isoformat(),
            source="manual",
        )

        # 중복 ID 체크
        existing_ids = {d.id for d in self._drugs}
        if drug.id in existing_ids:
            # suffix 추가
            counter = 2
            while f"{drug.id}-{counter}" in existing_ids:
                counter += 1
            drug.id = f"{drug.id}-{counter}"

        self._drugs.append(drug)
        self.save_to_json()
        logger.info(f"약품 추가: {drug.trade_name} (id={drug.id})")
        return drug

    def remove_drug(self, drug_id: str) -> bool:
        """약품 제거."""
        before = len(self._drugs)
        self._drugs = [d for d in self._drugs if d.id != drug_id]
        if len(self._drugs) < before:
            self.save_to_json()
            return True
        return False

    # ─── 검색어 생성 ──────────────────────────────────────

    def generate_search_keywords(self, drug: TargetDrug) -> dict[str, Any]:
        """소스별 검색에 사용할 키워드 생성.

        Returns:
            {
                "trade_name": "Rosumeg Combigel",
                "ingredient_full": "Rosuvastatin 5mg or 10mg + Omega-3-...",
                "ingredient_parts": ["Rosuvastatin 5mg or 10mg", "Omega-3-Acid Ethyl Esters 90 1g"],
                "ingredient_names": ["Rosuvastatin", "Omega-3-Acid Ethyl Esters 90"],
                "strength_normalized": "5 mg + 1000 mg" or None,
                "dosage_form_normalized": "capsule" or None,
            }
        """
        # 성분을 "+" 기준으로 분리
        parts = [p.strip() for p in drug.ingredient.split("+") if p.strip()]

        # 각 성분에서 용량 제거 → 성분명만 추출
        names = []
        seen_names: set[str] = set()
        for part in parts:
            # "Rosuvastatin 5mg or 10mg" → "Rosuvastatin"
            # "Omega-3-Acid Ethyl Esters 90 2g" → "Omega-3-Acid Ethyl Esters 90"
            # "Fluticasone Propionate 250μg or 500μg" → "Fluticasone Propionate"
            # "Gadobutrol 604.72 mg" → "Gadobutrol"
            name = re.sub(
                r"\s*\d+(?:[.,]\d+)?\s*(?:mg|mcg|g|ml|iu|%|μg|µg)\b.*$",
                "",
                part,
                flags=re.IGNORECASE,
            ).strip()
            if name and name.lower() not in seen_names:
                seen_names.add(name.lower())
                names.append(name)

        return {
            "trade_name": drug.trade_name,
            "ingredient_full": drug.ingredient,
            "ingredient_parts": parts,
            "ingredient_names": names,
            "strength_normalized": normalize_strength(drug.strength),
            "dosage_form_normalized": normalize_dosage_form(drug.dosage_form),
        }

    def to_dict_list(self) -> list[dict]:
        """약품 목록을 dict 리스트로 반환 (API 응답용)."""
        return [asdict(d) for d in self.list_drugs()]
